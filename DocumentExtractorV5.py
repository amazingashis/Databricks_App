import re
import ast
import copy
import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, List
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from collections import defaultdict

class NameSubstitutor(ast.NodeTransformer):
    """NodeTransformer to replace specified variable Names with Constant values."""
    def __init__(self, replacements):
        self.replacements = replacements

    def visit_Name(self, node):
        if node.id in self.replacements:
            return ast.copy_location(ast.Constant(value=self.replacements[node.id]), node)
        return node

class MappingExtractor(ast.NodeVisitor):
    def __init__(self):
        self.mappings = []
        self.join_details = []
        self.table_list_definitions = {}
        self.column_list_definitions = {}
        self.join_assignment_definitions = {}
        self.join_counts = {}
        self.dataframe_mappings = {}
        self.dataframe_to_table = {}
        self.dataframe_lineages = {}
        self.dq_rules = {}
        self.processed_withcolumn_nodes = set()
        self.current_table_name = None
        self.target_tables = None

    def _remove_function_names(self, long_text):
        """
           This function will strip known functional wrappers and return
           only the base column name, e.g., removing `explode` or `col`.
           Returns a LIST of column names, not a string.
        """
        # pattern = r"coalesce\(\s*'([^']*)'\s*,\s*lit\('NA'\)\s*\)|explode\(\s*col\('([^']+)'\)\s*\)"

        # coalesce_pattern = r"coalesce\(\s*'\"['\"]\s*,\s*lit\(['\"][^'\"]+['\"]\)\s*\)"
        # explode_pattern = r"explode(?:_outer)?\(\s*(?:F\.)?col\(\s*'\"['\"]\s*\)\s*\)"

        coalesce_pattern = r"""coalesce\(\s*(?:F\.)?col\(\s*'"['"]\s*\)\s*,.*?\)"""
        explode_pattern = r"""explode(?:_outer)?\(\s*(?:F\.)?col\(\s*'"['"]\s*\)\s*\)"""

        columns = []
        # columns += re.findall(pattern, long_text)

        columns.extend(re.findall(coalesce_pattern, long_text))
        columns.extend(re.findall(explode_pattern, long_text))

        # Deduplicate while preserving order
        seen = set()
        deduped = []
        for c in columns:
            if c not in seen:
                seen.add(c)
                deduped.append(c)
        return deduped

        # pattern = r"coalesce\(\s*'([^']*)'\s*,\s*lit\('NA'\)\s*\)|explode\(\s*col\('([^']+)'\)\s*\)"
        # matches = re.findall(pattern, long_text)
        # parameters = [p.strip("'") for match in matches for p in match if p]
        # return ', '.join(parameters)

    def _get_call_name(self, func):
        """Return the function name for both Name('col') and Attribute(Name('F'), 'col')."""
        if isinstance(func, ast.Name):
            return func.id
        if isinstance(func, ast.Attribute):
            return func.attr
        return None

    def _is_call(self, node, names: set):
        return isinstance(node, ast.Call) and self._get_call_name(node.func) in names

    def _slice_value(self, subscript):
        """Version-safe extractor for Subscript slice literal/identifier."""
        s = subscript.slice
        # Python 3.8: ast.Index, Python 3.9+: direct node
        if hasattr(ast, "Index") and isinstance(s, ast.Index):  # 3.8 style
            s = s.value
        if isinstance(s, ast.Constant):
            return s.value
        try:
            return ast.unparse(s)
        except Exception:
            return str(s)

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        bad = set(r'[]:*?/\'')
        clean = ''.join(ch for ch in name if ch not in bad)
        if len(clean) > 31:
            clean = clean[:31]
        return clean or "Sheet1"

    def try_unparse(self, node):
        try:
            return ast.unparse(node)
        except Exception:
            return ast.dump(node)
            # return str(node)

    def get_func_name(self, func_node):
        if isinstance(func_node, ast.Name):
            return func_node.id
        elif isinstance(func_node, ast.Attribute):
            return func_node.attr
        return None

    def get_base_df_or_method_name(self, node):
        """Get the base DataFrame or method name in a chain of calls or attribute access."""
        if isinstance(node, ast.Name):
            return node.id
        elif isinstance(node, ast.Attribute):
            return self.get_base_df_or_method_name(node.value)
        elif isinstance(node, ast.Call):
            if isinstance(node.func, ast.Attribute):
                return self.get_base_df_or_method_name(node.func.value)
            elif isinstance(node.func, ast.Name):
                return node.func.id
        return None

    def get_full_attribute_name(self, node):
        parts = []
        while isinstance(node, ast.Attribute):
            parts.append(node.attr)
            node = node.value
        if isinstance(node, ast.Name):
            parts.append(node.id)
        return ".".join(reversed(parts))

    def safe_extract_constant(self, args, index=0):
        """Safely extract the constant value from a node argument list."""
        if args and len(args) > index:
            arg = args[index]
            if isinstance(arg, ast.Constant):
                return arg.value
        return None

    def handle_alias(self, df_name, source_expr, target):
        transformation = self.try_unparse(source_expr)
        src_cols = self.extract_col_names(source_expr)
        src_table_cols, src_tables = self.map_columns_to_tables(src_cols)
        self.append_mapping(src_tables or df_name, src_table_cols, target, transformation)

    def extract_withcolumn_chain(self, node):
        """
        Recursively extract all withColumn calls from a chained expression.
        Returns a list of Call nodes in order (from base to last).
        """
        chain = []

        while isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute) and node.func.attr == 'withColumn':
                chain.insert(0, node)
                node = node.func.value
        return chain

    def expand_list_comprehension(self, node):
        """
        Expand a list comprehension AST node (or a Starred containing one) into a list of code strings.
        Returns a list of expanded expressions (as strings) or None if expansion isn't possible.
        """
        if isinstance(node, ast.ListComp):
            comp = node
        elif isinstance(node, ast.Starred) and isinstance(node.value, ast.ListComp):
            comp = node.value
        else:
            return None

        if len(comp.generators) != 1:
            return None

        gen = comp.generators[0]
        if isinstance(gen.target, ast.Name):
            var_names = [gen.target.id]
        elif isinstance(gen.target, ast.Tuple):
            var_names = [name.id for name in gen.target.elts if isinstance(name, ast.Name)]
        else:
            return None

        iter_values = None
        if isinstance(gen.iter, ast.Name):
            iter_name = gen.iter.id
            if iter_name in self.column_list_definitions:
                try:
                    iter_node = ast.List(elts=self.column_list_definitions[iter_name], ctx=ast.Load())
                    iter_values = ast.literal_eval(iter_node)
                except Exception:
                    iter_values = None
        elif isinstance(gen.iter, (ast.List, ast.Tuple)):
            try:
                iter_values = ast.literal_eval(gen.iter)
            except Exception:
                iter_values = None

        if iter_values is None:
            return None

        expanded_exprs = []
        for item in iter_values:
            values = item if isinstance(item, tuple) else (item,)
            if len(values) != len(var_names):
                continue

            replacements = {var: val for var, val in zip(var_names, values)}

            template_ast = copy.deepcopy(comp.elt)
            substituted_ast = NameSubstitutor(replacements).visit(template_ast)
            ast.fix_missing_locations(
                substituted_ast)

            code_str = ast.unparse(substituted_ast)
            expanded_exprs.append(code_str)
        return expanded_exprs

    def visit_Assign(self, node):
        if isinstance(node.targets[0], ast.Name):
            var_name = node.targets[0].id

            # if var_name in self.dataframe_to_table:
                # print(f'DataFrame Alias Set: {var_name} -> {self.dataframe_to_table[var_name]}')

            if var_name == 'dq_rule_string' and isinstance(node.value, (ast.Constant, ast.Expr)):
                if isinstance(node.value, ast.Constant):
                    dq_rule_value = node.value.value
                else:
                    dq_rule_value = self.try_unparse(node.value) or ""
                self.dq_rules = self.extract_dq_rules(dq_rule_value)

            if isinstance(node.value, ast.ListComp):
                expanded_exprs = self.expand_list_comprehension(node.value)
                if expanded_exprs:
                    expanded_nodes = [ast.parse(expr, mode='eval').body for expr in expanded_exprs]
                    self.column_list_definitions[var_name] = expanded_nodes
                    return

            if isinstance(node.value, ast.Call):
                call_expr = node.value
                if isinstance(call_expr.func, ast.Name):
                    func_name = call_expr.func.id
                    if func_name == 'table':
                        table_name = node.value.args[0].value
                        self.dataframe_to_table[var_name] = table_name

                elif isinstance(call_expr.func, ast.Attribute):
                    func_name = call_expr.func.attr
                    if func_name in ['select', 'withColumn', 'withColumns']:
                        self.process_list_variable(var_name, call_expr.args)
                    if func_name in ['groupBy', 'agg']:
                        base_df_name = self.get_base_df_or_method_name(call_expr.func.value)
                        if base_df_name in self.dataframe_to_table:
                            self.dataframe_to_table[var_name] = self.dataframe_to_table[base_df_name]
                            self.process_aggregation_expr(var_name, node.value)

            if isinstance(node.value, (ast.List, ast.Tuple)):
                self.column_list_definitions[var_name] = node.value.elts
                table_elements = [
                    elt.value for elt in node.value.elts if isinstance(elt, ast.Constant)
                ]
                self.table_list_definitions[var_name] = table_elements
                # print(f"Captured table list: {var_name} -> {table_elements}")
                # print("Table list definitions:", self.table_list_definitions)

            elif isinstance(node.value, ast.Call) and isinstance(node.value.func, ast.Name):
                if node.value.func.id == 'create_dataframe_from_table' and len(node.value.args) >= 3 and isinstance(node.value.args[2], ast.Name):
                    table_var_name = node.value.args[2].id
                    tables = self.table_list_definitions.get(table_var_name, [])
                    self.dataframe_mappings[var_name] = tables
                    if tables:
                        self.current_table_name = tables[0]
                        if len(tables) > 1:
                            print(f"Dataframe {var_name} created with multiple tables: {tables}")
                        else:
                            print(f"Dataframe {var_name} created with single table: {tables}")

            elif isinstance(node.value, ast.Call) and isinstance(node.value.func, ast.Attribute):
                func_base = node.value.func.value
                if isinstance(func_base, ast.Subscript):
                    subscript_val = func_base.value.id if isinstance(func_base.value, ast.Name) else None
                    slice_name = self._slice_value(func_base)
                    if subscript_val == "dm_df" and slice_name:
                        self.dataframe_to_table[var_name] = slice_name

                if node.value.func.attr == 'table' and node.value.args:
                    table_name = node.value.args[0].value if isinstance(node.value.args[0], ast.Constant) else None
                    if table_name:
                        self.dataframe_to_table[var_name] = table_name
                        # self.current_table_name = table_name
                        print(f"Mapping table name {table_name} to DataFrame {var_name}")

                if node.value.func.attr in ['filter', 'select', 'groupBy']:
                    base_df_name = node.value.func.value.id if isinstance(node.value.func.value, ast.Name) else None
                    if base_df_name in self.dataframe_to_table:
                        self.dataframe_to_table[var_name] = self.dataframe_to_table[base_df_name]
                        print(f'Filter applied. DataFrame Alias Set: {var_name} -> {self.dataframe_to_table[var_name]}')


            if isinstance(node.value, ast.Subscript):
                subscript_value = node.value.value
                # self.dataframe_lineages[var_name] = self.get_source_name(node.value)
                subscript_val = node.value.value.id if isinstance(node.value.value, ast.Name) else None
                slice_name = self._slice_value(node.value)
                if subscript_val == "dm_df" and slice_name:
                    self.dataframe_to_table[var_name] = slice_name

                if isinstance(subscript_value, ast.Name):
                    src_df = subscript_value.id
                    slice_name = self._slice_value(node.value)
                    if isinstance(slice_name, str):
                        self.dataframe_mappings[var_name] = [slice_name]
                        self.dataframe_to_table[var_name] = slice_name
                        self.current_table_name = slice_name


            if var_name == 'target_table' and isinstance(node.value, ast.JoinedStr):
                    formatted_parts = []
                    for elt in node.value.values:
                        if isinstance(elt, ast.Constant):
                            formatted_parts.append(elt.value.lstrip('.'))
                        # elif isinstance(elt, ast.FormattedValue):
                        #     if isinstance(elt.value, ast.Name):
                        #         formatted_parts.append(f"<{elt.value.id}>")

                    if formatted_parts:
                        self.target_tables = ''. join(formatted_parts)

        self.generic_visit(node)

    def visit_Call(self, node):
        try:
            if isinstance(node.func, ast.Name) and node.func.id == 'create_dataframe_from_table':
                if len(node.args) >= 3 and isinstance(node.args[2], ast.Name):
                    table_list_var = node.args[2].id
                    table_names = self.table_list_definitions.get(table_list_var, [])
                    if table_names:
                        self.current_table_name = table_names[0]
                    # print(f"Using tables from list: {self.current_table_name}")

            if isinstance(node.func, ast.Attribute):
                method = node.func.attr
                df_name = self.extract_df_name(node)

                # df_name = node.func.value.id if isinstance(node.func.value, ast.Name) else "Unknown"

                if df_name:
                    lineage_source = self.dataframe_lineages.get(df_name, df_name)
                    self.dataframe_lineages[df_name] = lineage_source

                if df_name in self.dataframe_mappings:
                    table_names = self.dataframe_mappings[df_name]
                    if table_names:
                        self.current_table_name = table_names[0]

                if df_name in self.dataframe_to_table:
                    self.current_table_name = self.dataframe_to_table[df_name]

                if method == "alias" and node.args and isinstance(node.args[0], ast.Constant):
                    alias = node.args[0].value
                    self.dataframe_to_table[alias] = self.dataframe_to_table.get(df_name, df_name)

                if method == 'select':
                    df_name = node.func.value.id if isinstance(node.func.value, ast.Name) else None
                    table_name = self.dataframe_to_table.get(df_name, df_name)

                    for arg in node.args:
                        if isinstance(arg, ast.Starred) and isinstance(arg.value, ast.ListComp):
                            expanded_exprs = self.expand_list_comprehension(arg.value)
                            if expanded_exprs:
                                for expr_code in expanded_exprs:
                                    expr_ast = ast.parse(expr_code, mode='eval').body
                                    self.process_select_expr(expr_ast, table_name)
                                continue

                        if isinstance(arg, ast.Starred):
                            starred_expr = arg.value
                            if isinstance(starred_expr, ast.Name) and starred_expr.id in self.column_list_definitions:
                                self.process_list_variable(table_name, self.column_list_definitions[starred_expr.id])
                            else:
                                transformation = self.try_unparse(arg)
                                self.append_mapping(table_name, "", "", transformation)

                        elif isinstance(arg, ast.Name) and arg.id in self.column_list_definitions:
                            self.process_list_variable(table_name or self.current_table_name, self.column_list_definitions[arg.id])
                            continue

                        else:
                            self.process_select_expr(arg, table_name)
                    return

                elif method == 'withColumn':
                    chain = self.extract_withcolumn_chain(node)
                    df_var = self.extract_df_name(chain[0])
                    default_table = self.dataframe_to_table.get(df_var, self.current_table_name)

                    for withcol_node in chain:
                        node_str = self.try_unparse(withcol_node)
                        if node_str in self.processed_withcolumn_nodes:
                            continue
                        self.processed_withcolumn_nodes.add(node_str)

                        self.process_withColumn(default_table, withcol_node)
                    # self.process_withColumn(df_name, node)

                elif method == 'withColumns':
                    self.process_withColumns(self.current_table_name, node)

                elif method == 'join':
                    self.process_join(node, df_name)

            self.generic_visit(node)
        except Exception as e:
            print(f"Error processing node: {self.try_unparse(node)}\nError: {e}")

    def get_full_source_name(self, node):
        if isinstance(node, ast.Attribute):
            return self.get_full_source_name(node.value) + "." + node.attr
        elif isinstance(node, ast.Name):
            return node.id
        return None

    def get_source_name(self, node):
        """Helper to extract the source name for a dataframe"""
        if isinstance(node, ast.Subscript):
            return f"{node.value.id}[{self.try_unparse(node.slice)}]"
        return "Unknown"

    def normalize_table_name(self, table_name):
        return table_name.split('.')[-1]

    def extract_dq_rules(self, dq_rule_string):
        lines = dq_rule_string.strip().split("\n")
        dq_rules = {}

        column_name_regex = re.compile(r'(?:(\w+)\s)|(?:(\w+)\s*\()')
        function_regex = re.compile(r'\w+\((\w+)\)')

        for line in lines:
            stripped_line = line.strip()
            if line.startswith("    "):
                rule = stripped_line

                match = function_regex.search(stripped_line)
                if match:
                    column_name = match.group(1)
                else:
                    column_match = column_name_regex.match(stripped_line)
                    column_name = column_match.group(1) if column_match else None

                if column_name:
                    dq_rules[column_name] = dq_rules.get(column_name, []) + [rule]

        for key, rules in dq_rules.items():
            dq_rules[key] = "; ".join(rules)

        return dq_rules

    def extract_array_fields_from_comprehension(self, expr):
        """Extract fields from an array of structs created with list comprehension."""
        fields = []

        if not isinstance(expr, ast.Call) or not expr.args:
            return fields

        comprehension = expr.args[0]
        if isinstance(comprehension, ast.Starred):
            comprehension = comprehension.value

        if isinstance(comprehension, ast.ListComp):
            struct_expr = comprehension.elt
            if isinstance(struct_expr, ast.Call) and self._get_call_name(struct_expr.func) == "struct":
                for field_expr in struct_expr.args:
                    if isinstance(field_expr, ast.Call) and self._get_call_name(field_expr.func) == "alias":
                        alias_name = field_expr.args[0].value if field_expr.args and isinstance(field_expr.args[0], ast.Constant) else None
                        transformation = self.try_unparse(field_expr)

                        fields.append((alias_name, transformation))
        else:
            fields.extend(self.extract_array_fields(expr))

        return fields

    def process_list_variable(self, table_name, list_elts):
        for elt in list_elts:
            self.process_select_expr(elt, table_name)

    def process_select_expr(self, expr, table_name="select()"):
        """Handles col(...).alias(...) inside .select()"""

        table_name = table_name or self.current_table_name

        if isinstance(expr, ast.Call):
            call_name = self._get_call_name(expr.func)

            if call_name == "alias":
                array_fields = []
                source_expr = expr.func.value

                # if isinstance(source_expr, ast.Call) and self._get_call_name(source_expr.func) == "array":
                #     array_fields = self.extract_array_fields_from_comprehension(source_expr)

                if isinstance(source_expr, ast.Call) and self._get_call_name(source_expr.func) == "array":
                    expanded_exprs = self.expand_list_comprehension(source_expr.args[0])
                    if expanded_exprs:
                        for struct_code in expanded_exprs:
                            struct_ast = ast.parse(struct_code, mode='eval').body
                            if isinstance(struct_ast, ast.Call) and self._get_call_name(struct_ast.func) == "struct":
                                for field_expr in struct_ast.args:
                                    if (isinstance(field_expr, ast.Call) and
                                            self._get_call_name(field_expr.func) == "alias"):
                                        alias_name = (field_expr.args[0].value
                                                      if field_expr.args and isinstance(field_expr.args[0], ast.Constant)
                                                      else None)
                                        transformation = self.try_unparse(field_expr)
                                        array_fields.append((alias_name, transformation))
                    else:
                        array_fields = self.extract_array_fields(source_expr) if hasattr(self, 'extract_array_fields') else []

                target = expr.args[0].value if expr.args and isinstance(expr.args[0], ast.Constant) else None
                transformation = self.try_unparse(expr)

                src_cols = self.extract_col_names(source_expr)
                src_table_cols, src_tables = self.map_columns_to_tables(src_cols)
                if not src_table_cols:
                    single = self.extract_col_name(source_expr)
                    if single:
                        src_table_cols, src_tables = self.map_columns_to_tables([single])
                src_table_display = src_tables or (table_name or "unknown")

                self.append_mapping(src_table_display, src_table_cols or "", target, transformation, array_fields)
                return

            elif call_name == "explode":
                target = self.try_unparse(expr.args[-1]) if expr.args else None
                source_expr = expr.args[0] if expr.args else None
                col_name = self.extract_col_name(source_expr)
                transformation = self.try_unparse(expr)
                src_table_cols, src_tables = self.map_columns_to_tables([col_name], default_table=table_name)
                self.append_mapping(src_tables, src_table_cols, target, transformation)
                return

            elif call_name == "col":
                col_name = expr.args[0].value if expr.args and isinstance(expr.args[0], ast.Constant) else None
                if col_name:
                    src_table_cols, src_tables = self.map_columns_to_tables([col_name], default_table=table_name)
                    self.append_mapping(src_tables, col_name, None, "col(\"%s\")" % col_name)
                return

        self.generic_visit(expr)
        # Non-alias calls like bare col("x") in select(...)
        # if isinstance(expr, ast.Call):
        #     source_col = self.extract_col_name(expr)
        #     transformation = self.try_unparse(expr)
        #     names = self._remove_function_names(transformation)
        #     col_name = source_col or (names[0] if names else "")
        #     self.append_mapping(table_name or "unknown", col_name, source_col or "", transformation)

    def extract_array_fields(self, expr):
        """Extracts field names from an array of structs"""
        fields = []
        if isinstance(expr, ast.Call) and self.get_func_name(expr.func) == "array":
            for struct_expr in expr.args:
                if struct_expr.func.id == "struct":
                    for sub_expr in struct_expr.args:
                        if isinstance(sub_expr, ast.Call) and isinstance(sub_expr.func, ast.Attribute) and sub_expr.func.attr == "alias":
                            alias_name = sub_expr.args[0].value if sub_expr.args else ""
                            source_expr = sub_expr.func.value
                            source_col = self.try_unparse(source_expr)
                            fields.append((alias_name, source_col))
        return fields

    def extract_df_name(self, node):
        """
        Recursively extract the base DataFrame name from a nested withColumn chain.
        """
        while isinstance(node, ast.Call):
            if isinstance(node.func, ast.Attribute):
                node = node.func.value
            else:
                break
        if isinstance(node, ast.Attribute):
            node = node.value
        return node.id if isinstance(node, ast.Name) else None

    def process_withColumn(self, table_name, call_node):
        if len(call_node.args) >= 2:
            # target = self.safe_extract_constant(call_node.args, 0)
            target = call_node.args[0]
            expr = call_node.args[1]

            if isinstance(expr, ast.Call) and isinstance(expr.func, ast.Attribute):
                base_expr = expr.func.value
            else:
                base_expr = expr

            array_fields = self.extract_array_fields(expr)

            if isinstance(target, ast.Constant):
                col_name, transformation, source_tables = self.process_transformation_and_extraction(
                    base_expr, default_table=table_name
                )
                src_table_display = source_tables or (table_name or self.current_table_name)
                self.append_mapping(src_table_display, col_name, target.value, transformation, array_fields)

    def process_withColumns(self, table_name, call_node):
        if call_node.args and isinstance(call_node.args[0], ast.Dict):

            df_var = call_node.func.value.id if isinstance(call_node.func.value, ast.Name) else None
            default_table = self.dataframe_to_table.get(df_var, table_name or self.current_table_name)

            arg = call_node.args[0]
            for key, val in zip(arg.keys, arg.values):
                if isinstance(key, ast.Constant):
                    col_name, transformation, source_tables = self.process_transformation_and_extraction(
                        val, default_table=default_table
                    )
                    src_table_display = source_tables or (table_name or self.current_table_name)
                    self.append_mapping(src_table_display, col_name, key.value, transformation)

    def process_join(self, node, current_df_name):
        try:
            left_df = self.resolve_dataframe_name(node.func.value) or current_df_name

            right_arg = node.args[0] if node.args else None
            right_df = self.resolve_dataframe_name(right_arg)

            on_expr = None
            how_val = None
            if len(node.args) >= 2:
                on_expr = node.args[1]
            if len(node.args) >= 3 and isinstance(node.args[2], ast.Constant):
                how_val = node.args[2].value

            for kw in node.keywords or []:
                if kw.arg == "on":
                    on_expr = kw.value
                elif kw.arg == "how" and isinstance(kw.value, ast.Constant):
                    how_val = kw.value.value

            join_type = (how_val or "left").upper()
            join_condition = self.try_unparse(on_expr) if on_expr is not None else "unknown"
            remarks = ""
            # remarks = f"This is from: {self.join_assignment_definitions[join_condition]}"

            # for condition in self.join_assignment_definitions:
            #     if condition == join_condition:
            #         remarks = f"From: {self.join_assignment_definitions[join_condition]}"

            for df_var, table_name in self.dataframe_to_table.items():
                join_condition = join_condition.replace(f"{df_var}.", f"{table_name}.")

            primary_table = self.dataframe_to_table.get(left_df, left_df)
            secondary_table = self.dataframe_to_table.get(right_df, right_df)

            self.join_details.append({
                "Primary Table": primary_table,
                "Secondary Table": secondary_table,
                "Join Type": join_type,
                "Join Condition": join_condition,
                "Remarks": remarks
            })
        except Exception as e:
            print(f"Error processing join node: {ast.dump(node)}. Exception: {e}")

        #     join_type = node.args[2].value if len(node.args) > 2 and isinstance(node.args[2], ast.Constant) else "left"
        #     join_condition = self.try_unparse(node.args[1]) if len(node.args) > 1 else "unknown"
        #
        #     col_pattern = r'col\("\'__([\w]+)["\']\)'
        #     for match in re.findall(col_pattern, join_condition):
        #         df_var, col_name = match
        #         table_name = self.dataframe_to_table.get(df_var, df_var)
        #         join_condition = re.sub(
        #             f'col\\(["\']{df_var}__{col_name}["\']\\)',
        #             f'{table_name}.{col_name}',
        #             join_condition
        #         )
        #
        #     for df_var, table_name in self.dataframe_to_table.items():
        #         join_condition = join_condition.replace(f"{df_var}.", f"{table_name}.")
        #
        #     primary_table = self.dataframe_to_table.get(left_df, left_df)
        #     secondary_table = self.dataframe_to_table.get(right_df, right_df)
        #
        #     # primary_table = self.get_full_source_name(left_df_alias)
        #     # secondary_table = self.get_full_source_name(right_df_alias)
        #
        #     self.join_details.append({
        #         "Primary Table": primary_table,
        #         "Secondary Table": secondary_table,
        #         "Join Type": join_type.upper(),
        #         "Join Condition": join_condition
        #     })
        #     print(f"Processed Join: {primary_table} joined with {secondary_table}, type: {join_type}")
        # except Exception as e:
        #     print(f"Error processing join node: {ast.dump(node)}. Exception: {e}")

    def resolve_dataframe_name(self, node):
        """Resolve the full DataFrame name from a potentially complex AST node."""
        if isinstance(node, ast.Name):
            return node.id
        elif isinstance(node, ast.Attribute):
            return self.resolve_dataframe_name(node.value)
        elif isinstance(node, ast.Call):
            if self._get_call_name(node.func) in ['select', 'filter', 'join']:
                return self.resolve_dataframe_name(node.func.value)
        elif isinstance(node, ast.Subscript):
            return self.resolve_dataframe_name(node.value)
        return "unknown"

    def extract_alias(self, node):
        try:
            if isinstance(node, ast.Name):
                return node.id
            elif isinstance(node, ast.Attribute):
                return node.attr if isinstance(node.value, ast.Name) else self.extract_alias(node.value)
        except Exception as e:
            print(f"Error extracting alias node: {ast.dump(node)}. Exception: {e}")
        return "unknown"

    def process_insert_into(self, node):
        if len(node.args) > 0 and isinstance(node.args[0], ast.Constant):
            self.target_tables = node.args[0].value

    def process_aggregation_expr(self, df_name, expr):
        resolved_table_name = self.dataframe_to_table.get(df_name, df_name)
        if isinstance(expr, ast.Call):
            func_name = self._get_call_name(expr.func)

            if func_name in ['agg', 'groupBy']:
                for call_node in expr.args:
                    if isinstance(call_node, ast.Call) and self._get_call_name(call_node.func) == "alias":
                        primary_transformation = self.try_unparse(call_node)
                        nested_fields = self.extract_struct_fields(call_node.func.value.args[0])
                        source_columns, _, _ = self.process_transformation_and_extraction(call_node, resolved_table_name)

                        array_fields = [(alias, source) for alias, source in nested_fields]

                        self.append_mapping(
                            resolved_table_name, source_columns, call_node.args[0].value, primary_transformation, array_fields
                        )

    def extract_struct_fields(self, struct_expr):
        """Extracts individual fields from a struct expression"""
        fields = []
        if isinstance(struct_expr, ast.Call):
            func_name = self._get_call_name(struct_expr.func)
            if func_name == "struct":
                for sub_expr in struct_expr.args:
                    if isinstance(sub_expr, ast.Call) and isinstance(sub_expr.func, ast.Attribute) and sub_expr.func.attr == "alias":
                        alias_name = sub_expr.args[0].value if sub_expr.args else ""
                        source_expr = sub_expr.func.value
                        source_col = self.try_unparse(source_expr)
                        fields.append((alias_name, source_col))
            else:
                for arg in struct_expr.args:
                    fields.extend(self.extract_struct_fields(arg))
        return fields

    def process_transformation_and_extraction(self, expr, default_table=None):
        transformation = self.try_unparse(expr).replace("\n", "")
        col_names = self.extract_col_names(expr)

        if not col_names and transformation:
            col_names = self._remove_function_names(transformation)

        if isinstance(col_names, str):
            col_names = [c.strip() for c in col_names.split(",") if c.strip()]
        elif not isinstance(col_names, (list, tuple)):
            col_names = []

        col_names = [c for c in col_names if c and self.is_valid_column_name(c)]

        source_table_col_names, source_tables = self.map_columns_to_tables(
            col_names, default_table=self.dataframe_to_table.get(default_table, default_table)
        )
        transformation = transformation.replace(r'\n', '').replace('  ', '')
        return source_table_col_names, transformation, source_tables

    def map_columns_to_tables(self, col_names, default_table=None):
        """Maps DataFrame column references to table.column format"""
        table_col_names = []
        tables = set()

        default = default_table or self.current_table_name or "unknown"

        for col in col_names:
            if not isinstance(col, str) or not col:
                continue
            if '.' in col:
                parts = col.split('.', 1)
                df_var = parts[0]
                column = parts[1] if len(parts) > 1 else ""
                # df_var, column = col.split('.', 1)
                full_table_name = self.dataframe_to_table.get(df_var, df_var)

            else:
                full_table_name = default
                column = col
                # df_var, column = (self.current_table_name, col)
            # full_table_name = self.dataframe_to_table.get(df_var, df_var)
            table_name = full_table_name

            if '.' in full_table_name:
                _, table_name = full_table_name.split(".")

            if column and self.is_valid_column_name(column):
                table_col_names.append(f"{table_name}.{column}")
            # else:
            #     table_col_names.append(table_name)

            tables.add(full_table_name)

        return ", ".join(table_col_names), ", ".join(sorted(tables))

    def append_mapping(self, table_name, col_name, target, transformation, array_fields=None):
        if array_fields is None:
            array_fields = []

        if not col_name.strip():
            table_name = ""

        resolved_table_name = self.dataframe_to_table.get(table_name, table_name)
        if resolved_table_name == "Unknown" and table_name in self.dataframe_to_table:
            resolved_table_name = self.dataframe_to_table[table_name]

        mapping = {
            "Source Table": resolved_table_name,
            "Source Column(s)": col_name,
            "Target Column(s)": target,
            "Array Field": [(alias, trans) for alias, trans in array_fields],
            "Transformation": transformation
        }
        self.mappings.append(mapping)

    def _attribute_chain(self, node):
        """Return ['df', 'col', 'subfield', ...] from a nested Attribute node."""
        parts = []
        while isinstance(node, ast.Attribute):
            parts.append(node.attr)
            node = node.value
        if isinstance(node, ast.Name):
            parts.append(node.id)
        return list(reversed(parts))

    def extract_col_name(self, node):
        """
        Extract a single source column name for simple cases like:
        - col("...") or F.col("...").
        - df.colName(.subField...)
        Returns the literal column string when unambiguous; otherwise None.
        """
        if isinstance(node, ast.Call):
            func = node.func
            fname = func.id if isinstance(func, ast.Name) else (
                func.attr if isinstance(func, ast.Attribute) else None)
            if fname == "col" and node.args and isinstance(node.args[0], ast.Constant) and isinstance(node.args[0].value, str):
                return node.args[0].value

        if isinstance(node, ast.Attribute):
            parts = self._attribute_chain(node)
            if len(parts) >= 2:
                return ".".join(parts)
        return None

    def is_valid_column_name(self, column_name):
        """
            Determine if the string is a valid column name.
            Filters out known method names, function names, date formats, and invalid symbols.
            """
        if not isinstance(column_name, str) or not column_name:
            return False

        unwanted_names = {
            'alias', 'isNull', 'isNotNull', 'isin', 'otherwise', 'when', 'between',
            'startswith', 'endswith', 'contains', 'cast', 'substr', 'rlike', 'like',
            'over', 'desc', 'asc', 'asc_nulls_first', 'asc_nulls_last', 'desc_nulls_first',
            'desc_nulls_last', 'getItem',
            "coalesce", "substring", "upper", "lower", "trim", "ltrim", "rtrim",
            "regexp_replace", "regexp_extract", "concat", "concat_ws", "collect_set",
            "length", "date_format", "to_date", "to_timestamp", "from_unixtime",
            "unix_timestamp", "nvl", "ifnull", "explode", "explode_outer", "posexplode",
            "posexplode_outer", "element_at", "array", "struct",
            "yyyy-MM-dd", "MM-dd-yyyy", "dd-MM-yyyy", "yyyyMMdd"
        }

        invalid_symbols = set("-:<>/\\|&=+*^%$#@!~`")
        if column_name in unwanted_names:
            return False
        if any(char in column_name for char in invalid_symbols):
            return False

        return column_name.isidentifier()

        # return column_name.isidentifier() and not any(symbol in column_name for symbol in "-: ")

    def extract_col_names(self, expr):
        """Extract all col("...") used in an expression"""
        cols = set()

        METHOD_ATTRS = {
            'alias', 'isNull', 'isNotNull', 'isin', 'otherwise', 'when', 'between',
            'startswith', 'endswith', 'contains', 'cast', 'substr', 'rlike', 'like',
            'over', 'desc', 'asc', 'asc_nulls_first', 'asc_nulls_last', 'desc_nulls_first',
            'desc_nulls_last', 'getItem'
        }

        def get_call_name(func):
            if isinstance(func, ast.Name):
                return func.id
            if isinstance(func, ast.Attribute):
                return func.attr
            return None

        FUNC_NAMES = {
            "col",
            "coalesce", "substring", "upper", "lower", "trim", "ltrim", "rtrim",
            "regexp_replace", "regexp_extract", "concat", "concat_ws", "collect_set",
            "when", "otherwise", "length", "date_format", "to_date",
            "to_timestamp", "from_unixtime", "unix_timestamp", "nvl", "ifnull",
            "explode", "explode_outer", "posexplode", "posexplode_outer",
            "element_at", "array", "struct"
        }

        DATE_FORMATS = {"yyyy-MM-dd", "MM-dd-yyyy", "dd-MM-yyyy"}

        class ColVisitor(ast.NodeVisitor):
            def __init__(self, outer):
                # super().__init__()
                self.outer = outer

            def visit_Call(self, node):
                # fname = get_call_name(node.func)

                # if fname == "col" and node.args and isinstance(node.args[0], ast.Constant):
                #     arg0 = node.args[0]
                #     if isinstance(arg0.value, str):
                #         cols.add(arg0.value)

                func_name = self.outer._get_call_name(node.func)
                if func_name == "col" and node.args:
                    arg = node.args[0]
                    if isinstance(arg, ast.Constant) and isinstance(arg.value, str):
                        column_name = arg.value
                        if self.outer.is_valid_column_name(column_name):
                            cols.add(column_name)

                elif func_name in FUNC_NAMES:
                    for arg in node.args:
                        if isinstance(arg, ast.Constant) and isinstance(arg.value, str):
                            val = arg.value
                            if (
                                    self.outer.is_valid_column_name(val)
                                    and val not in METHOD_ATTRS
                                    and val not in FUNC_NAMES
                                    and val not in DATE_FORMATS
                                    and not re.match(r"[<>&].*?\\", val)
                            ):
                                cols.add(val)

                if isinstance(node.func, ast.Name) and node.func.id in FUNC_NAMES:
                    for arg in node.args:
                        if isinstance(arg, ast.Constant) and isinstance(arg.value, str) and arg.value not in DATE_FORMATS:
                            cols.add(arg.value)

                    if node.args and isinstance(node.args[0], ast.Constant) and isinstance(node.args[0].value, str):
                        cols.add(node.args[0].value)

                for arg in node.args:
                    self.visit(arg)
                for kw in node.keywords or []:
                    self.visit(kw.value)

                self.generic_visit(node)

            def visit_Attribute(self, node):
                parts = self.outer._attribute_chain(node)
                if parts and self.outer.is_valid_column_name(parts[0]):
                    cols.add(".".join(parts))

                if len(parts) >= 2:
                    head = parts[0]
                    tail = parts[1:]
                    # keep everything up to the first method
                    cut = len(tail)
                    for i, p in enumerate(tail):
                        if p in METHOD_ATTRS:
                            cut = i
                            break
                    if cut > 0:
                        col_path = tail[:cut]
                        cols.add(head + "." + ".".join(col_path))
                self.generic_visit(node)

        ColVisitor(self).visit(expr)
        return sorted(cols)

        #     def visit_Attribute(self, node):
        #         full_attr = self.parent.get_full_attribute_name(node)
        #         if full_attr != 'alias':
        #             cols.add(full_attr)
        #         self.generic_visit(node)
        #
        # ColVisitor(self).visit(expr)
        # return sorted(cols)

        # return ", ".join(sorted(cols)) if cols else None


def modify_excel(file_path, target_table_name, join_details_data):
    from openpyxl.styles import Alignment, Font

    wb = load_workbook(file_path)

    if "Join Details" in wb.sheetnames:
        del wb["Join Details"]
    join_ws = wb.create_sheet("Join Details")

    headers = ["Primary Table", "Secondary Table", "Join Type", "Join Condition", "Remarks"]
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    for col_idx, h in enumerate(headers, 1):
        cell = join_ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold_font
        cell.alignment = center_alignment

    for r_idx, row in enumerate(join_details_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            join_ws.cell(row=r_idx, column=c_idx, value=val)

    from openpyxl.styles import Alignment
    join_ws.freeze_panes = "A2"
    widths = [25, 25, 12, 80, 40]
    for i, w in enumerate(widths, start=1):
        join_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in join_ws.iter_rows(min_row=2, max_row=join_ws.max_row, min_col=1, max_col=join_ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(file_path)

def extract_dataframe_lineages(code: str) -> Dict[str, List[str]]:

    class Tracker(ast.NodeVisitor):
        def __init__(self):
            self.lineages = defaultdict(list)

        def visit_Assign(self, node):
            target = node.targets[0].id if isinstance(node.targets[0], ast.Name) else None
            sources = self._extract_sources(node.value)
            if target and sources:
                filtered = [s for s in sources if not s.startswith("dm_df") and s != target]
                if filtered:
                    self.lineages[target].extend(filtered)
            self.generic_visit(node)

        def _extract_sources(self, node):
            sources = []
            if isinstance(node, ast.Call):
                if isinstance(node.func, ast.Attribute) and node.func.attr == "table":
                    if node.args and isinstance(node.args[0], ast.Constant):
                        sources.append(node.args[0].value)
                        return sources
                base = self._resolve_base(node)
                if base:
                    sources.append(base)
                if isinstance(node.func, ast.Attribute) and node.func.attr == "join":
                    left = self._resolve_base(node.func.value)
                    right = self._resolve_base(node.args[0])
                    if left: sources.append(left)
                    if right: sources.append(right)
            elif isinstance(node, ast.Subscript):
                base = self._resolve_base(node.value)
                if base:
                    sources.append(f"{base}[...]")
            return list(set(sources))

        def _resolve_base(self, node):
            while isinstance(node, ast.Call):
                if isinstance(node.func, ast.Attribute):
                    node = node.func.value
                else:
                    break
            return node.id if isinstance(node, ast.Name) else None

    tree = ast.parse(code)
    tracker = Tracker()
    tracker.visit(tree)
    return dict(tracker.lineages)

def generate_mapping(file_path, output_excel_prefix="mapping_output", output_folder="mapping_outputs"):
    source = Path(file_path).read_text(encoding="utf-8")
    tree = ast.parse(source)
    # print(tree.body[10]._fields)

    extractor = MappingExtractor()
    extractor.visit(tree)

    # extractor.dataframe_lineages = extract_dataframe_lineages(source)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    raw_target = extractor.target_tables or 'unknown'
    safe_target = MappingExtractor._sanitize_sheet_name(
        raw_target.replace('.', '_').replace('<', '').replace('>', '')
    )
    target_table_name = safe_target
    output_excel_name = f"{output_excel_prefix}_{target_table_name}_{timestamp}.xlsx"

    output_folder_path = Path(output_folder)
    output_folder_path.mkdir(parents=True, exist_ok=True)
    output_excel = output_folder_path / output_excel_name


    new_mappings = []
    for mapping in extractor.mappings:
        full_source_table = mapping["Source Table"]
        # source_table = full_source_table.split(".")[-1]
        target_col = mapping["Target Column(s)"]
        dq_rule = extractor.dq_rules.get(target_col, "")
        mapping["DQ Rules"] = dq_rule

        new_mappings.append({
            "Source Table": mapping["Source Table"],
            "Source Column(s)": mapping["Source Column(s)"],
            "Target Column(s)": mapping["Target Column(s)"],
            "Array Field": "",
            "Transformation": mapping["Transformation"],
            "DQ Rules": dq_rule
        })

        array_fields = mapping["Array Field"]

        if array_fields:
            for  alias, transformation in array_fields:
                new_mappings.append({
                    "Source Table": "",
                    "Source Column(s)": "",
                    "Target Column(s)": "",
                    "Array Field": alias,
                    "Transformation": transformation if transformation else "",
                    "DQ Rules": ""
                })

    df = pd.DataFrame(new_mappings)
    df = df[["Source Table", "Source Column(s)", "Target Column(s)", "Array Field", "Transformation", "DQ Rules"]]

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        workbook = writer.book
        worksheet = workbook.active
        worksheet.freeze_panes = worksheet['A2']
        if target_table_name:
            worksheet.title = MappingExtractor._sanitize_sheet_name(target_table_name)

        column_widths = {1: 25, 2: 45, 3: 25, 4: 30, 5: 50, 6: 30}
        for col_idx, width in column_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    if extractor.join_details:
        join_details_data = [
            (detail["Primary Table"], detail["Secondary Table"], detail["Join Type"], detail["Join Condition"], detail["Remarks"])
            for detail in extractor.join_details
        ]
        modify_excel(output_excel, target_table_name, join_details_data)

    print(f"Mapping document generated: {output_excel}")

# === USAGE ===
if __name__ == "__main__":
    generate_mapping("load_silver_edw_member_service_provider.py")

    # To generate from a folder

    # folder_path = Path('mappings')
    #
    # for file in folder_path.iterdir(
    #     if file.is_file():
    #         print(f"Reading file: {file}")
    #         generate_mapping(file)

    # OR

    # files = ['member_and_enrollment_table.py', 'load_silver_echo_provider_service_org_relationship.py']
    # for f in files:
    #     generate_mapping(f)
